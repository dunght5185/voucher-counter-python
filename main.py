# import module
import openpyxl
from openpyxl import Workbook


inputFile = "Data Sales update 22.09.2023.xlsx"
outputFile = "Output - " + inputFile


# load excel with its path
wrkbk = openpyxl.load_workbook(inputFile)
sh = wrkbk.worksheets[0]

workbook = Workbook()
worksheet = workbook.active
worksheet.append(["Time", "Shop Name", "Voucher Code", "Number of uses"])

row_count = sh.max_row
column_count = sh.max_column

i = 0
tempI = 0
maxCount = 1
count = 1

currentBuyDate = ''
currentShopName = ''
currentVoucherCode = ''
currentCostPrice = ''

tempBuyDate = ''
tempShopName = ''
tempVoucherCode = ''
tempCostPrice = ''

nextBuyDate = ''
nextShopName = ''
nextVoucherCode = ''
nextCostPrice = ''
rowNum = row_count

while i < rowNum:

    currentBuyDate = sh.cell(row=i + 1, column=2).value
    currentShopName = sh.cell(row=i + 1, column=13).value
    currentVoucherCode = sh.cell(row=i + 1, column=24).value
    currentCostPrice = sh.cell(row=i + 1, column=26).value

    nextBuyDate = sh.cell(row=i + 2, column=2).value
    nextShopName = sh.cell(row=i + 2, column=13).value
    nextVoucherCode = sh.cell(row=i + 2, column=24).value
    nextCostPrice = sh.cell(row=i + 2, column=26).value


    if(currentCostPrice != 0):        
        if(currentBuyDate == nextBuyDate and currentShopName == nextShopName and currentVoucherCode == nextVoucherCode and nextCostPrice != 0):
            count += 1
        else:
            print(currentBuyDate, currentShopName, currentVoucherCode, maxCount)
            worksheet.append([currentBuyDate, currentShopName, currentVoucherCode, maxCount])
            count = 1
    else:
        print(".....................")
    
    # print(currentShopName)
    maxCount = count
    i = i + 1
workbook.save(outputFile)
