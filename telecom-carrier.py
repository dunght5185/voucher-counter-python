# import module
import openpyxl
from openpyxl import Workbook


inputFile = "KHMM092023 (25.09).xlsx"
outputFile = "Output - " + inputFile


# load excel with its path
wrkbk = openpyxl.load_workbook(inputFile)
sh = wrkbk.worksheets[0]

workbook = Workbook()
worksheet = workbook.active
worksheet.append(["STT", "Tên khách hàng", "Địa chỉ", "Tên người liên hệ","SĐT", "Nhà mạng"])

row_count = sh.max_row
column_count = sh.max_column

i = 1
rowNum = row_count
rowNum = 271
tempTenKH = ''
tempDiaChi = ''
tempNguoiLienHe = ''
tempSoDT = ''
tempNhaMang = ''

def timNhaMang(sdt):
    match sdt:
        case "086" | "096" | "097" | "098" | "039" | "038" | "037" | "036" | "035" | "034" | "033" | "032":
            return "Viettel"
        case "070" | "079" | "077" | "076" | "078" | "089" | "090" | "093":
            return "Mobiphone"
        case "091" | "094" | "088" | "083" | "084" | "085" | "081" | "082":
            return "Vinaphone"
        case "092" | "052" | "056" | "058":
            return "Vietnamobile"
        case "059" | " 099":
            return "Gmobile"
        case default:
            return "something"
  
 


while i < rowNum:
    
    tempTenKH = sh.cell(row=i, column=3).value
    tempDiaChi = sh.cell(row=i, column=4).value
    tempNguoiLienHe = sh.cell(row=i, column=5).value
    tempSoDT = str(sh.cell(row=i, column=6).value)
    tempNhaMang = timNhaMang(tempSoDT[0:3])
    print(i, tempTenKH, tempDiaChi, tempNguoiLienHe, tempSoDT, tempNhaMang)

    worksheet.append([i, tempTenKH, tempDiaChi, tempNguoiLienHe, tempSoDT, tempNhaMang])

    i = i + 1
workbook.save(outputFile)
